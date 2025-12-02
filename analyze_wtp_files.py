#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script d'analyse des fichiers WTP (Water Treatment Plant) pour le cahier des charges
"""

import sys
import io

# Forcer UTF-8 pour l'encodage
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("openpyxl n'est pas installé. Installation en cours...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from pathlib import Path
import json

def analyze_excel_wtp(file_path):
    """Analyse un fichier Excel WTP"""
    print(f"\n{'='*80}")
    print(f"Analyse du fichier: {file_path.name}")
    print(f"{'='*80}\n")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        print(f"Nombre de feuilles: {len(sheet_names)}")
        print(f"Noms des feuilles: {sheet_names}\n")
        
        analysis = {
            'file_name': file_path.name,
            'file_path': str(file_path),
            'number_of_sheets': len(sheet_names),
            'sheets': [],
            'type': 'WTP (Water Treatment Plant)'
        }
        
        for sheet_name in sheet_names:
            print(f"\n--- Feuille: {sheet_name} ---")
            
            try:
                ws = wb[sheet_name]
                max_row = ws.max_row
                max_col = ws.max_column
                
                print(f"Dimensions réelles: {max_row} lignes x {max_col} colonnes")
                
                # Analyser les premières lignes pour détecter la structure
                header_candidates = []
                sample_data = []
                indicators = []
                
                for row_idx in range(1, min(30, max_row + 1)):
                    row_values = []
                    non_null_count = 0
                    
                    for col_idx in range(1, min(max_col + 1, 50)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        
                        if value is not None:
                            try:
                                str_value = str(value).strip()
                                if str_value:
                                    non_null_count += 1
                                    row_values.append(str_value[:80])
                                    # Détecter les indicateurs/colonnes importantes
                                    if any(keyword in str_value.upper() for keyword in ['DATE', 'TIME', 'VALEUR', 'VALUE', 'PARAMETRE', 'PARAMETER', 'CONTROLE', 'CHECK', 'PRODUCTION', 'CONSOMMATION', 'FLOW', 'DEBIT', 'PRESSURE', 'PRESSION', 'PH', 'TDS', 'CONDUCTIVITY', 'CONDUCTIVITE']):
                                        indicators.append(str_value)
                            except:
                                pass
                        else:
                            row_values.append("")
                    
                    if non_null_count > max_col * 0.2:
                        header_candidates.append(row_idx - 1)
                        if len(sample_data) < 15:
                            sample_data.append({
                                'row': row_idx,
                                'values': row_values[:20]
                            })
                
                if header_candidates:
                    print(f"\nCandidats pour les en-têtes (lignes): {header_candidates[:10]}")
                    
                    print(f"\nPremières lignes de données (échantillon):")
                    for sample in sample_data[:8]:
                        print(f"  Ligne {sample['row']}: {sample['values'][:8]}")
                
                if indicators:
                    print(f"\nIndicateurs/Colonnes importantes détectés: {indicators[:15]}")
                
                # Détecter les cellules fusionnées
                merged_cells = []
                try:
                    if hasattr(ws, 'merged_cells') and ws.merged_cells:
                        merged_cells = list(ws.merged_cells.ranges)
                except AttributeError:
                    pass
                
                if merged_cells:
                    print(f"\nCellules fusionnées détectées: {len(merged_cells)}")
                    print(f"Premières cellules fusionnées: {[str(mc) for mc in merged_cells[:5]]}")
                
                # Détecter les formules
                formula_count = 0
                formula_samples = []
                for row in ws.iter_rows(max_row=min(200, max_row), max_col=min(max_col, 50)):
                    for cell in row:
                        if cell.data_type == 'f':
                            formula_count += 1
                            if len(formula_samples) < 5:
                                try:
                                    formula_samples.append({
                                        'cell': cell.coordinate,
                                        'formula': str(cell.value)[:100]
                                    })
                                except:
                                    pass
                
                if formula_count > 0:
                    print(f"\nFormules détectées (dans les 200 premières lignes): {formula_count}")
                    if formula_samples:
                        print("Exemples de formules:")
                        for f in formula_samples[:3]:
                            print(f"  {f['cell']}: {f['formula'][:60]}...")
                
                # Analyser les données pour détecter les paramètres de traitement d'eau
                water_params = []
                for col_idx in range(1, min(max_col + 1, 30)):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    sample_values = []
                    
                    for row_idx in range(1, min(50, max_row + 1)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            try:
                                val_str = str(cell.value)[:50]
                                sample_values.append(val_str)
                                # Détecter des paramètres de traitement d'eau
                                if any(param in val_str.upper() for param in ['PH', 'TDS', 'CONDUCTIVITY', 'TURBIDITY', 'CHLORINE', 'FLOW', 'PRESSURE', 'RO', 'UF', 'PERMEAT', 'REJECT', 'PERMEATE']):
                                    if val_str not in water_params:
                                        water_params.append(val_str)
                            except:
                                pass
                            if len(sample_values) >= 5:
                                break
                
                if water_params:
                    print(f"\nParamètres de traitement d'eau détectés: {water_params[:10]}")
                
                sheet_analysis = {
                    'name': sheet_name,
                    'max_rows': max_row,
                    'max_columns': max_col,
                    'merged_cells_count': len(merged_cells),
                    'has_formulas': formula_count > 0,
                    'formula_count_sample': formula_count,
                    'header_candidates': header_candidates[:10],
                    'sample_data': sample_data[:10],
                    'indicators': indicators[:20],
                    'water_parameters': water_params[:20]
                }
                analysis['sheets'].append(sheet_analysis)
                
            except Exception as e:
                print(f"Erreur lors de l'analyse de la feuille {sheet_name}: {str(e)}")
                analysis['sheets'].append({
                    'name': sheet_name,
                    'error': str(e)
                })
        
        wb.close()
        return analysis
        
    except Exception as e:
        print(f"Erreur lors de l'analyse du fichier {file_path}: {str(e)}")
        return {
            'file_name': file_path.name,
            'error': str(e)
        }

def main():
    base_path = Path("docs/Re_ Documents du WTP")
    
    # Fichiers Excel à analyser
    excel_files = [
        "PS3-ID17 SYNTHESE DES CONSOMMATIONS.xlsx",
        "PS3-ID15 RAPPORT DE PRODUCTION.xlsx",
        "PS3-ID16 RO log sheet for operators - 30.xls",
        "check-lists WTP.xlsx"
    ]
    
    all_analyses = []
    
    for file_name in excel_files:
        file_path = base_path / file_name
        if file_path.exists():
            try:
                analysis = analyze_excel_wtp(file_path)
                all_analyses.append(analysis)
            except Exception as e:
                print(f"Erreur lors de l'analyse de {file_name}: {str(e)}")
                all_analyses.append({
                    'file_name': file_name,
                    'error': str(e)
                })
        else:
            print(f"Fichier non trouvé: {file_path}")
    
    # Analyser aussi les fichiers Word pour obtenir les titres et structure
    word_files = [
        "PS3-ID14 CONTROLE TRAITEMENT EAU NETTOYAGE DES EQUIPEMENTS - 30.docx",
        "PS3-ID13 CONTROLE TRAITEMENT EAU FILTRATION - 30.docx",
        "PS3-ID12 CONTROLE TRAITEMENT EAU PERMEAT - 30.docx",
        "PS3-ID11 CONTROLE TRAITEMENT EAU OSMOSE INVERSE - 30 - Copy.docx",
        "PS3-ID10 CONTROLE TRAITEMENT DE L'EAU - 30 - Copy.docx",
        "PS3-ID09 CONTROLE TRAITEMENT DE L'EAU BRUTE - 30 - Copy.docx"
    ]
    
    print(f"\n\n{'='*80}")
    print("Fichiers Word détectés (non analysés automatiquement):")
    for word_file in word_files:
        file_path = base_path / word_file
        if file_path.exists():
            print(f"  - {word_file}")
            all_analyses.append({
                'file_name': word_file,
                'file_path': str(file_path),
                'type': 'Word Document (Checklist)',
                'note': 'Fichier Word de checklist de contrôle traitement d\'eau - nécessite analyse manuelle'
            })
    
    # Sauvegarder l'analyse dans un fichier JSON
    output_file = Path("wtp_files_analysis.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_analyses, f, ensure_ascii=False, indent=2)
    
    print(f"\n\n{'='*80}")
    print("Analyse complète terminée. Résultats sauvegardés dans wtp_files_analysis.json")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()

